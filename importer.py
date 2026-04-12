import openpyxl
from datetime import datetime
from db import get_connection, init_db


def parse_date(value):
    """Convert various date formats to ISO YYYY-MM-DD string, or None."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    s = str(value).strip()
    for fmt in ("%d %b %Y", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None  # unrecognised format — store nothing rather than bad data


def load_manifest(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path)
    conn = get_connection()

    with conn:
        _seed_ssr_codes(conn, wb)
        flight_key = _insert_flight(conn, wb)
        _insert_passengers(conn, wb, flight_key)

    conn.close()
    print(f"Loaded manifest from {xlsx_path}")


def _seed_ssr_codes(conn, wb):
    """Populate ssr_codes from the Fare & Code Key sheet (SSR rows only)."""
    ws = wb["Fare & Code Key"]
    ssr_section = False
    for row in ws.iter_rows(values_only=True):
        if row[0] == "SSR CODES":
            ssr_section = True
            continue
        if ssr_section and row[0] and row[1]:
            conn.execute(
                "INSERT OR IGNORE INTO ssr_codes (code, description) VALUES (?, ?)",
                (str(row[0]).strip(), str(row[1]).strip()),
            )


def _insert_flight(conn, wb):
    """Read flight metadata from Manifest Cover and insert into flights."""
    ws = wb["Manifest Cover"]
    meta = {}
    for row in ws.iter_rows(values_only=True):
        if row[0] and row[1]:
            meta[str(row[0]).strip()] = str(row[1]).strip()

    flight_no   = meta.get("Flight Number")
    flight_date = parse_date(meta.get("Flight Date"))
    # Origin/destination cells include verbose descriptions (e.g. "SFO – San Francisco..."),
    # so extract just the 3-letter IATA code before the first space or dash.
    origin      = meta.get("Origin", "").split()[0]
    destination = meta.get("Destination", "").split()[0]
    conn.execute(
        """INSERT INTO flights (flight_no, flight_date, origin, destination, operator)
           VALUES (?, ?, ?, ?, ?)""",
        (flight_no, flight_date, origin, destination, meta.get("Operator")),
    )
    return (flight_no, flight_date, origin, destination)


def _insert_passengers(conn, wb, flight_key):
    """Insert all passengers from the Passenger Manifest sheet."""
    flight_no, flight_date, origin, destination = flight_key
    ws = wb["Passenger Manifest"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {name: idx for idx, name in enumerate(headers)}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[col["No."]] is None:
            continue  # skip empty trailing rows

        conn.execute(
            """INSERT INTO passengers (
                flight_no, flight_date, origin, destination,
                record_locator, last_name, first_name, title,
                gender, seat, cabin_class, fare_class, e_ticket_no,
                ssr_codes, notes
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                flight_no, flight_date, origin, destination,
                row[col["Record Locator"]],
                row[col["Last Name"]],
                row[col["First Name"]],
                row[col["Title"]],
                row[col["Gender"]],
                row[col["Seat"]],
                row[col["Cabin Class"]],
                row[col["Fare Class"]],
                row[col["E-Ticket No."]],
                row[col["SSR Codes"]],
                row[col["Notes"]],
            ),
        )


if __name__ == "__main__":
    import sys
    import os

    xlsx_path = sys.argv[1] if len(sys.argv) > 1 else "AB_123_Passenger_Manifest_Updated.xlsx"
    if not os.path.exists(xlsx_path):
        print(f"File not found: {xlsx_path}")
        sys.exit(1)

    init_db()
    load_manifest(xlsx_path)

    # Quick sanity check
    from db import get_connection
    conn = get_connection()
    pax_count = conn.execute("SELECT COUNT(*) FROM passengers").fetchone()[0]
    ssr_count = conn.execute("SELECT COUNT(*) FROM passengers WHERE ssr_codes IS NOT NULL").fetchone()[0]
    print(f"  {pax_count} passengers, {ssr_count} with SSR codes")
    conn.close()
